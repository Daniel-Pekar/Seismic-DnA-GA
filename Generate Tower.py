import os
import win32com.client
import openpyxl
import random
from openpyxl import *

#from geneticalgorithm.py###############################
class Gene:
    def __init__(self,ws,name="",lower=0,upper=0,value=0,cell_col = 0):
        self.cell_col = cell_col
        self.name = ws['K' + str(self.cell_col)].value
        self.lower = ws['L' + str(self.cell_col)].value
        self.upper = ws['M' + str(self.cell_col)].value
        self.value = value

    def generate_initial_values(self,ws):
        self.value = random.uniform(self.lower,self.upper)

class Chromosome:
    def __init__(self,len,genes = [],fitness = 0, chance = 0):
        self.len = len
        self.genes = genes
        self.fitness = fitness
        self.selectionchance = chance

    def create_Chromosome(self,worksheet):
        self.genes = [Gene(ws = worksheet, cell_col = i + 4) for i in range(self.len)]
        for gene in self.genes:
            gene.generate_initial_values(ws)

####################################################

class Node:
    def __init__(self, name, x, y, z, x_gene_name = 'na', y_gene_name = 'na', z_gene_name = 'na'):
        self.name = name
        self.x = x
        self.y = y
        self.z = z
        self.x_gene_name = x_gene_name
        self.y_gene_name = y_gene_name
        self.z_gene_name = z_gene_name


def multiply_by_gene(var_to_check, gene_to_mult_by, range_to_mult_low, range_to_mult_high):
    for current_gene_num in range(range_to_mult_low, range_to_mult_high):
        


def find_value_dep_on_gene(gene_name, chromosome):
    print(gene_name)
    if '-' in gene_name:
        gene_name_pos = gene_name.replace('-', '')
        for i in range(len(chromosome)):
            if chromosome[i].name == gene_name_pos:
                value = chromosome[i].value * -1
                break
    else:
        for i in range(len(chromosome)):
            if chromosome[i].name == gene_name:
                value = chromosome[i].value
                break
    return value

#note: put in the row/column indices as strings


def get_nodes(nodes_name_col, nodes_x_col, nodes_y_col, nodes_z_col, nodes_start_row, chromosome):
    wb = load_workbook('Setup.xlsx')
    ws = wb.active
    all_nodes = []
    #assign values to nodes
    current_row = nodes_start_row
    while ws[nodes_name_col + str(current_row)].value is not None:
        node_name = ws[nodes_name_col + str(current_row)].value
        node_x = ws[nodes_x_col + str(current_row)].value
        node_y = ws[nodes_y_col + str(current_row)].value
        node_z = ws[nodes_z_col + str(current_row)].value
        #check if the x or y locations are dependent on the value of a gene
        node_x_loc = 'na'
        node_y_loc = 'na'
        node_z_loc = 'na'
        if type(node_x) != int:
            node_x_loc = node_x
            node_x = find_value_dep_on_gene(node_x, chromosome)
        if type(node_y) != int:
            node_y_loc = node_y
            node_y = find_value_dep_on_gene(node_y, chromosome)
        if type(node_z) != int:
            node_z_loc = node_z
            node_z = find_value_dep_on_gene(node_z, chromosome)
        all_nodes.append(Node(node_name, node_x, node_y, node_z, node_x_loc, node_y_loc, node_z_loc))
        current_row = current_row + 1
    return all_nodes

wb = load_workbook('Setup.xlsx')
ws = wb.active

TestChromosome = Chromosome(len = 23)
TestChromosome.create_Chromosome(ws)

AllNodes = get_nodes('B', 'F', 'G', 'H', 4, TestChromosome.genes)


for i in range(len(AllNodes)):
    print(AllNodes[i].name)
    print(AllNodes[i].x)
    print(AllNodes[i].y)
    print(AllNodes[i].z)