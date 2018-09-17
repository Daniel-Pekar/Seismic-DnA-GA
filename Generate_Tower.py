import os
import win32com.client
import openpyxl
import random
from openpyxl import *
import re
import time
import GeneticAlgorithm
from GeneticAlgorithm import*


class Node:
    def __init__(self, name, x, y, z, lumped_mass, x_gene_name = 'na', y_gene_name = 'na', z_gene_name = 'na'):
        self.name = name
        self.x = x
        self.y = y
        self.z = z
        self.lumped_mass = lumped_mass
        self.x_gene_name = x_gene_name
        self.y_gene_name = y_gene_name
        self.z_gene_name = z_gene_name

class Member:
    def __init__(self, start_node, end_node, type, prop, thickness):
        self.start_node = start_node
        self.end_node = end_node
        self.type = type
        self.prop = prop
        self.thickness = thickness
        self.name = start_node + ' to ' + end_node + ' (' + type + ')'


def find_value_dep_on_gene(gene_name, range_to_mult_by_gene, gene_to_mult_by, chromosome):
    if '-' in gene_name:
        gene_name_pos = gene_name.replace('-', '')
        for i in range(len(chromosome)):
            if chromosome[i].name == gene_name_pos:
                value = chromosome[i].value * -1
                if multiply_by_gene(gene_name, range_to_mult_by_gene):
                    value = value * chromosome[gene_to_mult_by - 1].value
                break
    else:
        for i in range(len(chromosome)):
            if chromosome[i].name == gene_name:
                value = chromosome[i].value
                if multiply_by_gene(gene_name, range_to_mult_by_gene):
                    value = value * chromosome[gene_to_mult_by - 1].value
                break
    return value


def multiply_by_gene(var_to_check, range_to_mult):
    var_num_pos = var_to_check.replace('-', '')
    var_num = var_num_pos.replace('V', '')
    for current_gene_num in range_to_mult:
        if current_gene_num == int(var_num):
            return True
    return False


def get_nodes(nodes_name_col, nodes_x_col, nodes_y_col, nodes_z_col, nodes_mass_col, nodes_start_row, range_to_mult_by_gene, gene_to_mult_by, chromosome, ws):
    print('Reading nodes from Excel workbook...')
    all_nodes = []
    #assign values to nodes
    current_row = nodes_start_row
    while ws[nodes_name_col + str(current_row)].value is not None:
        node_name = ws[nodes_name_col + str(current_row)].value
        node_x = ws[nodes_x_col + str(current_row)].value
        node_y = ws[nodes_y_col + str(current_row)].value
        node_z = ws[nodes_z_col + str(current_row)].value
        node_mass = ws[nodes_mass_col + str(current_row)].value
        #check if the x or y locations are dependent on the value of a gene
        node_x_loc = 'na'
        node_y_loc = 'na'
        node_z_loc = 'na'
        if type(node_x) == str:
            node_x_loc = node_x
            node_x = find_value_dep_on_gene(node_x, range_to_mult_by_gene, gene_to_mult_by, chromosome)
        if type(node_y) == str:
            node_y_loc = node_y
            node_y = find_value_dep_on_gene(node_y, range_to_mult_by_gene, gene_to_mult_by, chromosome)
        if type(node_z) == str:
            node_z_loc = node_z
            node_z = find_value_dep_on_gene(node_z, range_to_mult_by_gene, gene_to_mult_by, chromosome)
        all_nodes.append(Node(node_name, node_x, node_y, node_z, node_mass, node_x_loc, node_y_loc, node_z_loc))
        current_row = current_row + 1
    return all_nodes


def get_members(start_node_col, end_node_col, member_type_col, member_prop_col, member_thickness_col, chromosome, start_row, ws):
    print('Reading members from Excel workbook...')
    all_members = []
    current_row = start_row
    while ws[start_node_col + str(current_row)].value is not None:
        start_node_name = ws[start_node_col + str(current_row)].value
        end_node_name = ws[end_node_col + str(current_row)].value
        member_type = ws[member_type_col + str(current_row)].value
        member_prop = ws[member_prop_col + str(current_row)].value
        member_thickness = None
        if member_prop == 'VARIABLE':
            member_thickness_var = ws[member_thickness_col + str(current_row)].value
            member_thickness = find_value_dep_on_gene(member_thickness_var, [], [], chromosome)
        all_members.append(Member(start_node_name, end_node_name, member_type, member_prop, member_thickness))
        current_row = current_row + 1
    return all_members


def build_tower(nodes, members, mat_props_cols, section_props_cols, start_row, ws, time_history, save_location):
    print('Initializing SAP2000 model...')
    # create SAP2000 object
    SapObject = win32com.client.Dispatch('SAP2000v15.SapObject')
    # start SAP2000
    SapObject.ApplicationStart()
    #create SapModel Object
    SapModel = SapObject.SapModel
    #initiaize model
    SapModel.InitializeNewModel()
    #create new blank model
    ret = SapModel.File.NewBlank()
    #set units for materials
    N_m_C = 10
    SapModel.SetPresentUnits(N_m_C)
    #define material properties
    print('Defining material properties...')
    for mat_prop_counter in range(len(mat_props_cols)):
        current_col = mat_props_cols[mat_prop_counter]
        mat_type = ws[current_col + str(start_row)].value
        mat_name = ws[current_col + str(start_row + 1)].value
        mat_E = ws[current_col + str(start_row + 2)].value
        mat_poisson = ws[current_col + str(start_row + 3)].value
        mat_thermal = ws[current_col + str(start_row + 4)].value
        mat_unit_wt = ws[current_col + str(start_row + 5)].value
        #create material type
        SapModel.PropMaterial.SetMaterial(mat_name, mat_type)
        #set isotropic material properties
        SapModel.PropMaterial.SetMPIsotropic(mat_name, mat_E, mat_poisson, mat_thermal)
        #set unit weight
        SapModel.PropMaterial.SetWeightAndMass(mat_name, 1, mat_unit_wt)
    #define section properties
    if section_props_cols != 'None':
        print('Defining section properties...')
        for sec_prop_counter in range(len(section_props_cols)):
            current_col = section_props_cols[sec_prop_counter]
            sec_name = ws[current_col + str(start_row)].value
            sec_mat = ws[current_col + str(start_row + 1)].value
            sec_shape = ws[current_col + str(start_row + 2)].value
            if 'SQUARE' == sec_shape:
                sec_width = ws[current_col + str(start_row + 3)].value
                sec_height = ws[current_col + str(start_row + 4)].value
                SapModel.PropFrame.SetRectangle(sec_name, sec_mat, sec_height, sec_width)
            elif 'CIRCULAR' == sec_shape:
                sec_dia = ws[current_col +str(start_row + 3)].value
                SapModel.PropFrame.SetCircle(sec_name, sec_mat, sec_dia)
            else:
                print('ERROR: Define a section shape in the setup workbook!')
    #create nodes
    print('Creating nodes...')
    total_bad_nodes = 0
    for current_node_counter in range(len(nodes)):
        current_node_name = nodes[current_node_counter].name
        current_node_x = nodes[current_node_counter].x
        current_node_y = nodes[current_node_counter].y
        current_node_z = nodes[current_node_counter].z
        current_node_weight = nodes[current_node_counter].lumped_mass
        assigned_node_name = ' '
        #set units to inches
        lb_in_F = 1
        SapModel.SetPresentUnits(lb_in_F)
        [ret, assigned_node_name] = SapModel.PointObj.AddCartesian(current_node_x, current_node_y, current_node_z, assigned_node_name, current_node_name)
        if ret != 0:
            print('ERROR creating node' + current_node_name)
            total_bad_nodes = total_bad_nodes + 1
        #set mass and loads, if necessary
        if current_node_weight != 0:
            N_m_C = 10
            SapModel.SetPresentUnits(N_m_C)
            ret = SapModel.PointObj.SetMassByWeight(current_node_name, [current_node_weight, 0, 0, 0, 0, 0], 0, True, False)
            if ret[0] != 0:
                print('ERROR setting lumped mass at node ' + current_node_name)
            ret = SapModel.PointObj.SetLoadForce(current_node_name, 'DEAD', [0, 0, current_node_weight, 0, 0, 0])
            if ret[0] != 0:
                print('ERROR setting joint weight at node ' + current_node_name)
        #set restraints, if necessary
        if current_node_z == 0:
            ret = SapModel.PointObj.SetRestraint(current_node_name, [True, True, True, True, True, True])
            if ret[0] != 0:
                print('ERROR setting joint restraint at node ' + current_node_name)
    print('Done creating nodes. Total bad nodes: ' + str(total_bad_nodes))
    #create members
    print('Creating members...')
    total_bad_members = 0
    for current_member_counter in range(len(members)):
        current_member_start_node = members[current_member_counter].start_node
        current_member_end_node = members[current_member_counter].end_node
        current_member_prop = members[current_member_counter].prop
        current_member_thickness = members[current_member_counter].thickness
        current_member_name = members[current_member_counter].name
        assigned_member_name = ''
        #set units to inches
        lb_in_F = 1
        SapModel.SetPresentUnits(lb_in_F)
        #if member thicknesses are variable
        if current_member_prop == 'VARIABLE':
            #create the section property for this thickness
            sec_name = str('Member' + str(current_member_counter))
            sec_mat = 'BALSA'
            sec_height = current_member_thickness
            sec_width = current_member_thickness
            SapModel.PropFrame.SetRectangle(sec_name, sec_mat, sec_height, sec_width)
            current_member_prop = sec_name
        [ret, assigned_member_name] = SapModel.FrameObj.AddByPoint(current_member_start_node, current_member_end_node, assigned_member_name, current_member_prop, current_member_name)
        if ret != 0:
            print('ERROR creating member ' + current_member_name)
            total_bad_members = total_bad_members + 1
        current_member_counter = current_member_counter + 1
    print('Done creating members. Total bad members: ' + str(total_bad_members))
    print('Creating load cases and combinations...')
    #Define time history function
    N_m_C = 10
    SapModel.SetPresentUnits(N_m_C)
    SapModel.Func.FuncTH.SetFromFile('GM', time_history, 1, 0, 1, 2, True)
    #Set the time history load case
    N_m_C = 10
    SapModel.SetPresentUnits(N_m_C)
    SapModel.LoadCases.ModHistLinear.SetCase('GM')
    SapModel.LoadCases.ModHistLinear.SetMotionType('GM', 1)
    SapModel.LoadCases.ModHistLinear.SetLoads('GM', 1, ['Accel'], ['U1'], ['GM'], [1], [1], [0], ['Global'], [0])
    SapModel.LoadCases.ModHistLinear.SetTimeStep('GM', 250, 0.1)
    #Create load combination
    SapModel.RespCombo.Add('DEAD + GM', 0)
    SapModel.RespCombo.SetCaseList('DEAD + GM', 0, 'DEAD', 1)
    SapModel.RespCombo.SetCaseList('DEAD + GM', 0, 'GM', 1)
    print('Finished constructing tower. Saving file...')
    #Save the model
    ret = SapModel.File.Save(save_location)
    if ret != 0:
        print('ERROR saving SAP2000 file')
    return SapObject


def get_sap_results(SapObject):
    # create SapModel Object
    SapModel = SapObject.SapModel
    #Run Analysis
    print('Computing...')
    SapModel.Analyze.RunAnalysis()
    print('Finished computing.')
    #Get RELATIVE acceleration from node 5-3-2
    SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
    SapModel.Results.Setup.SetComboSelectedForOutput('DEAD + GM', True)
    #set type to envelope
    SapModel.Results.Setup.SetOptionModalHist(1)
    #Get joint acceleration
    #Set units to metres
    N_m_C = 10
    SapModel.SetPresentUnits(N_m_C)
    g = 9.81
    ret = SapModel.Results.JointAccAbs('5-3-2', 0)
    max_and_min_acc = ret[7]
    max_pos_acc = max_and_min_acc[0]
    min_neg_acc = max_and_min_acc[1]
    if abs(max_pos_acc) >= abs(min_neg_acc):
        max_acc = abs(max_pos_acc)/g
    elif abs(min_neg_acc) >= abs(max_pos_acc):
        max_acc = abs(min_neg_acc)/g
    else:
        print('Could not find max acceleration')
    #Get joint displacement
    #Set units to millimetres
    N_mm_C = 9
    SapModel.SetPresentUnits(N_mm_C)
    ret = SapModel.Results.JointDispl('5-3-2', 0)
    max_and_min_disp = ret[7]
    max_pos_disp = max_and_min_disp[0]
    min_neg_disp = max_and_min_disp[1]
    if abs(max_pos_disp) >= abs(min_neg_disp):
        max_drift = abs(max_pos_acc)
    elif abs(min_neg_disp) >= abs(max_pos_disp):
        max_drift = abs(min_neg_disp)
    else:
        print('Could not find max drift')
    #Close SAP2000
    SapObject.ApplicationExit(True)
    return max_acc, max_drift


def get_excel_indices(ws, index_headings_col, index_values_col, index_start_row):
    excel_index = {}
    current_row = index_start_row
    while ws[index_headings_col + str(current_row)].value is not None:
        index_heading = ws[index_headings_col + str(current_row)].value
        index_value = ws[index_values_col + str(current_row)].value
        #check if there are multiple columns/values associated with a parameter
        if ',' in str(ws[index_values_col + str(current_row)].value):
            #create array of individual values
            index_value_array = ws[index_values_col + str(current_row)].value.split(',')
            contains_letters = False
            #check if the data should be formatted as integers or strings
            for i in range(len(index_value_array)):
                if re.search('[a-zA-Z]', index_value_array[i]):
                    contains_letters = True
            if not contains_letters:
                #format data as integers
                for j in range(len(index_value_array)):
                    index_value_array[j] = int(index_value_array[j])
            index_value = index_value_array
        #enter the new entry into the index
        excel_index[index_heading] = index_value
        current_row = current_row + 1
    return excel_index


def ga_CONSTRUCT(chromosome_genes, ws, excel_index, time_history, save_location):
    print('\nConstruct')
    print('----------------------------------')
    #variables for get nodes
    nodes_name_col = excel_index.get('Node name col')
    nodes_x_col = excel_index.get('Node x col')
    nodes_y_col = excel_index.get('Node y col')
    nodes_z_col = excel_index.get('Node z col')
    nodes_mass_col = excel_index.get('Node mass col')
    nodes_start_row = excel_index.get('Start row')
    range_to_mult_by_gene = excel_index.get('Variables to multiply')
    gene_to_mult_by = excel_index.get('Variable to multiply by')
    #variables for get members
    start_node_col = excel_index.get('Member start col')
    end_node_col = excel_index.get('Member end col')
    member_type_col = excel_index.get('Member type col')
    member_prop_col = excel_index.get('Member property col')
    member_thickness_col = excel_index.get('Member thickness col')
    member_start_row = excel_index.get('Start row')
    #variables for build tower
    mat_props_cols = excel_index.get('Material property defs')
    section_props_cols = excel_index.get('Section property defs')
    start_row = excel_index.get('Start row')
    all_nodes = get_nodes(nodes_name_col, nodes_x_col, nodes_y_col, nodes_z_col, nodes_mass_col, nodes_start_row, range_to_mult_by_gene, gene_to_mult_by, chromosome_genes, ws)
    all_members = get_members(start_node_col, end_node_col, member_type_col, member_prop_col, member_thickness_col, chromosome_genes, start_row, ws)
    SapObject = build_tower(all_nodes, all_members, mat_props_cols, section_props_cols, start_row, ws, time_history, save_location)
    return SapObject


def ga_ANALYZE(SapObject):
    print('\nAnalyze')
    print('----------------------------------')
    max_acc_and_drift = get_sap_results(SapObject)
    print('Max acceleration is: ' + str(max_acc_and_drift[0]) + ' g')
    print('Max drift is: ' + str(max_acc_and_drift[1]) + ' mm')
    return max_acc_and_drift
