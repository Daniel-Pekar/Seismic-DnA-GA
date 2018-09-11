import os
import win32com.client
import openpyxl
import random
from openpyxl import *
import bisect
from operator import attrgetter
import scipy
import numpy
from scipy.stats import norm

class Gene:
    def __init__(self,ws,name="",lower=0,upper=0,value=0,cell_col = 0):
        self.cell_col = cell_col
        self.name = ws['N' + str(self.cell_col)].value
        self.lower = ws['O' + str(self.cell_col)].value
        self.upper = ws['P' + str(self.cell_col)].value
        self.value = value

    def generate_initial_values(self,ws):
        self.value = random.uniform(self.lower,self.upper)

class Chromosome:
    def __init__(self,len,genes = [],fitness = 0):
        self.len = len
        self.genes = genes
        self.fitness = fitness

    def FABI(self, results):
        footprint = (self.genes[0].value*2)**2 #inches squared
        weight = 1.12 #lb
        design_life = 100 #years
        construction_cost = 2500000*(weight**2)+6*(10**6)
        land_cost = 35000 * footprint
        annual_building_cost = (land_cost + construction_cost) / design_life
        annual_revenue = 430300
        equipment_cost = 20000000
        return_period_1 = 50
        return_period_2 = 300
        max_disp = results[1] #mm
        apeak_1 = results[0] #g's
        xpeak_1 = 100*max_disp/1524 #% roof drift
        structural_damage_1 = scipy.stats.norm(1.5, 0.5).cdf(xpeak_1)
        equipment_damage_1 = scipy.stats.norm(1.75, 0.7).cdf(apeak_1)
        economic_loss_1 = structural_damage_1*construction_cost + equipment_damage_1*equipment_cost
        annual_economic_loss_1 = economic_loss_1/return_period_1
        structural_damage_2 = 0.5
        equipment_damage_2 = 0.5
        economic_loss_2 = structural_damage_2*construction_cost + equipment_damage_2*equipment_cost
        annual_economic_loss_2 = economic_loss_2/return_period_2
        annual_seismic_cost = annual_economic_loss_1 + annual_economic_loss_2
        fabi = annual_revenue - annual_building_cost - annual_seismic_cost
        return fabi

    def create_initial_chromosome(self,worksheet):
        self.genes = [Gene(ws = worksheet, cell_col = i + 4) for i in range(self.len)]
        for gene in self.genes:
            gene.generate_initial_values(worksheet)

    def mutation_uniform(self,prob):
        for i in range(self.len):
            if random.random()<prob:
                self.genes[i].value = random.uniform(self.genes[i].lower,self.genes[i].upper)

    def mutation_triangular(self,prob):
        for i in range(self.len):
            if random.random() < prob:
                self.genes[i].value = random.triangular(self.genes[i].lower, self.genes[i].upper)

    def mutation_min(self,prob):
        for i in range(self.len):
            if random.random() < prob:
                self.genes[i].value = random.uniform(self.genes[i].lower)

    def mutation_max(self,prob):
        for i in range(self.len):
            if random.random() < prob:
                self.genes[i].value = random.uniform(self.genes[i].upper)

class Population:
    def __init__(self, chromosomes = [],generation = 1,pop = 0,chromosomelen = 0):
        self.chromosomes = chromosomes
        self.generation = generation
        self.pop = pop
        self.chromlen = chromosomelen

    def create_initial_pop(self,ws):
        self.chromosomes = [Chromosome(len = self.chromlen) for i in range(self.pop)]
        for chromosome in self.chromosomes:
            chromosome.create_initial_chromosome(ws)

    def total_fit(self):
        return sum(chromosome.fitness for chromosome in self.chromosomes)

    def avg_fitness(self):
        return self.total_fit()/len(self.chromosomes)

    def max_fitness(self):
        maxf = 0
        for chromosome in self.chromosomes:
            if chromosome.fitness > maxf:
                maxf = chromosome.fitness
        return maxf

    def selection_elitism(self,num):
        self.chromosomes.sort(key=lambda x: x.fitness, reverse=True)
        parents = []
        for i in range(num):
            parents.append(self.chromosomes.pop(0))
        self.pop -= num
        return parents

    def selection_roulette(self,num_parents):
        self.chromosomes.sort(key=lambda x: x.fitness, reverse=False)
        parents = []
        fitness_list = []
        weights = []
        total_fit = 0
        for chromosome in self.chromosomes:
            fitness_list.append(chromosome.fitness)
        total_fit = sum(fitness_list)
        weights = [round(fit/total_fit, 4) for fit in fitness_list]
        weights[-1] = 0
        weights[-1] = 1-sum(weights)
        if weights[-1] < 0:
            weights[weights.index(max(weights))] -= 0.001
            weights[-1] += 0.001
        parents = numpy.random.choice(self.chromosomes, size = num_parents, replace = False, p = weights)
        return parents

    def selection_stochastic(self,num_parents):
        self.chromosomes.sort(key=lambda x: x.fitness, reverse=True)
        parents = []
        cumfit = 0
        fitness_list = [[], self.chromosomes]
        for chromosome in self.chromosomes:
            cumfit += chromosome.fitness
            fitness_list[0].append(cumfit)
        randnum = random.uniform(0, cumfit/num_parents)
        for i in range(num_parents):
            ind = bisect.bisect_left(fitness_list[0], randnum+(cumfit/num_parents)*(i))
            parents.append(fitness_list[1][ind])
        return parents

    def selection_tournament(self,num_parents,num_fighters):
        parents = []
        for i in range(num_parents):
            fighters = []
            for j in range(num_fighters):
                ind = random.randint(0,self.pop)
                fighters.append(self.chromosomes[ind])
                self.chromosomes.pop(ind)
            parents.append(max(fighters, key=attrgetter('fitness')))
        return parents

    def selection_rank(self,num_parents):
        self.chromosomes.sort(key=lambda x: x.fitness, reverse=True)
        weightings = []
        for i in range(self.pop):
            weightings.append(1/(i+1)) #1/(x+1) weightings for rank x
        parents = random.choices(self.chromosomes, weights = weightings, k = num_parents)
        return parents

    def crossover_npoint(self,n,parents):
        crossover_points = random.sample(range(1, self.chromlen), n)
        crossover_points.sort(key=lambda x: x, reverse=False)
        crossover_points.insert(0, 0)
        crossover_points.append(self.chromlen)
        children = [Chromosome(len = self.chromlen) for i in range(2)]
        origin = []
        for i in range(len(crossover_points) - 1):
            for j in range(crossover_points[i + 1] - crossover_points[i]):
                origin.append(i % 2)
        for i in range(len(origin)):
            children[0].genes.append(parents[origin[i]].genes[i])
            children[1].genes.append(parents[1 - origin[i]].genes[i])
        return children

    def crossover_randomflip(self,parents,num_children):
        children = []
        for i in range(num_children):
            children.append([])
            for j in range(self.chromlen):
                if random.randint(0,1) == 0:
                    children[-1].append(parents[0][j])
                else:
                    children[-1].append(parents[1][j])
        return children

    def crossover_flip(self,parents):
        children = [[],[]]
        for i in range(self.chromlen):
            children[0].append(parents[i%2][i])
            children[1].append(parents[1 - i%2][i])
        return children

    def crossover_triangle(self,parents,num_children):
        children = [[] for i in range(num_children)]
        for i in children:
            for j in range(self.chromlen):
                i.append(random.triangular(parents[0][j],parents[1][j]))
        return children

    def crossover_uniform(self,parents,num_children):
        children = [[] for i in range(num_children)]
        for i in children:
            for j in range(self.chromlen):
                i.append(random.uniform(parents[0][j],parents[1][j]))
        return children